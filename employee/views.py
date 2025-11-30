# employee/views.py
from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from .models import Employee
from .serializers import EmployeeSerializer

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER
from datetime import datetime

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# --- Custom Pagination ---
class StandardResultsSetPagination(PageNumberPagination):
    # Number of items per page
    page_size = 10
    # Optional: allows client to override page size up to a max
    page_size_query_param = "page_size"
    max_page_size = 100


# --- Employee ViewSet (CRUD, Filtering, Sorting, Pagination, Export) ---
class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    pagination_class = StandardResultsSetPagination

    # Filtering and Sorting Backend Setup
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]

    # Enable filtering by specific fields
    filterset_fields = ["department", "position", "hire_date"]

    # Enable searching across these fields
    search_fields = ["first_name", "last_name", "email", "department", "position"]

    # Enable ordering (sorting) by these fields
    ordering_fields = ["last_name", "hire_date", "salary"]

    # Default ordering
    ordering = ["last_name"]

    @action(detail=False, methods=["get"])
    def export_pdf(self, request):
        """
        Export employees to PDF
        URL: /api/employees/export_pdf/
        Supports same filters as main list (search, department, position, ordering)
        """
        # Get filtered queryset (respects filters from request)
        queryset = self.filter_queryset(self.get_queryset())

        # Create the HttpResponse object with PDF headers
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="employees_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        )

        # Create the PDF object using ReportLab
        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(letter),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=18,
        )

        # Container for the 'Flowable' objects
        elements = []

        # Add title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#1a1a1a"),
            spaceAfter=30,
            alignment=TA_CENTER,
        )
        title = Paragraph("Employee Directory", title_style)
        elements.append(title)

        # Add metadata
        info_style = ParagraphStyle(
            "Info",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.HexColor("#666666"),
            spaceAfter=20,
            alignment=TA_CENTER,
        )
        info_text = f"Generated on {datetime.now().strftime('%B %d, %Y at %H:%M')} | Total Employees: {queryset.count()}"
        elements.append(Paragraph(info_text, info_style))
        elements.append(Spacer(1, 12))

        # Create table data
        data = [["Name", "Email", "Position", "Department", "Hire Date", "Salary"]]

        for emp in queryset:
            data.append(
                [
                    f"{emp.first_name} {emp.last_name}",
                    emp.email,
                    emp.position,
                    emp.department,
                    emp.hire_date.strftime("%Y-%m-%d"),
                    f"₹{emp.salary:,.2f}",
                ]
            )

        # Create table
        table = Table(
            data,
            colWidths=[
                1.5 * inch,
                2 * inch,
                1.5 * inch,
                1.3 * inch,
                1 * inch,
                1 * inch,
            ],
        )

        # Add style to table
        table.setStyle(
            TableStyle(
                [
                    # Header styling
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 11),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("TOPPADDING", (0, 0), (-1, 0), 12),
                    # Body styling
                    ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                    ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 9),
                    ("ALIGN", (4, 1), (4, -1), "CENTER"),  # Center hire date
                    ("ALIGN", (5, 1), (5, -1), "RIGHT"),  # Right align salary
                    # Alternating row colors
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#F2F2F2")],
                    ),
                    # Grid
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("BOX", (0, 0), (-1, -1), 1, colors.black),
                    # Padding
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 1), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                ]
            )
        )

        elements.append(table)

        # Build PDF
        doc.build(elements)

        return response

    @action(detail=False, methods=["get"])
    def export_excel(self, request):
        """
        Export employees to Excel
        URL: /api/employees/export_excel/
        Supports same filters as main list (search, department, position, ordering)
        """
        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"

        # Define headers
        headers = [
            "ID",
            "First Name",
            "Last Name",
            "Email",
            "Phone",
            "Position",
            "Department",
            "Hire Date",
            "Salary",
            "Created At",
            "Updated At",
        ]
        ws.append(headers)

        # Style the header row
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data
        for emp in queryset:
            # Convert timezone-aware datetimes to naive datetimes for Excel
            created_at_naive = (
                emp.created_at.replace(tzinfo=None) if emp.created_at else None
            )
            updated_at_naive = (
                emp.updated_at.replace(tzinfo=None) if emp.updated_at else None
            )

            ws.append(
                [
                    emp.id,
                    emp.first_name,
                    emp.last_name,
                    emp.email,
                    emp.phone_number or "N/A",
                    emp.position,
                    emp.department,
                    emp.hire_date,
                    float(emp.salary),
                    created_at_naive,
                    updated_at_naive,
                ]
            )

        # Format salary column (column I, index 9)
        for row in range(2, ws.max_row + 1):
            salary_cell = ws.cell(row=row, column=9)
            salary_cell.number_format = "₹#,##0.00"
            salary_cell.alignment = Alignment(horizontal="right")

        # Format date columns
        date_columns = [8, 10, 11]  # Hire Date, Created At, Updated At
        for col in date_columns:
            for row in range(2, ws.max_row + 1):
                date_cell = ws.cell(row=row, column=col)
                date_cell.number_format = "YYYY-MM-DD HH:MM:SS"
                date_cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze the header row
        ws.freeze_panes = "A2"

        # Add filters to header row
        ws.auto_filter.ref = ws.dimensions

        # Create response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="employees_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        )

        wb.save(response)

        return response
